from openpyxl import *
import pyexcel

import requests
import os




class CSVModifer:

    def __init__(self):
        self.file = None
        self.worksheet = None
        self.workbook = None

    def set_file(self, file):
        self.file = file

    def get_file(self):
        return self.file

    def set_workbook(self, path):
        self.workbook = load_workbook(path)

    def set_worksheet(self, worksheet):
        self.worksheet = self.workbook[worksheet]

    def get_sheetnames(self):
        return self.workbook.sheetnames

    def index_of_element(self,inquiry):
        for x in range(len(self.workbook.sheetnames)):
            if self.workbook.sheetnames[x].lower() == inquiry.lower():
                return x
        return -1

    def element_of_index(self,inquiry):
        return self.workbook.sheetnames[inquiry]

    def load(self):
        for row in self.worksheet.values:
            for value in row:
                print(value)

